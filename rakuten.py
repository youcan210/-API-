import requests
import openpyxl as px
import pandas as pd
import fire
from datetime import datetime

class API:
    def __init__(self,url:str):
        self.url = url
        self.parameters = {}
    
    def get_parameters(self,parameters):
        self.parameters = parameters
        return self.parameters

    def get_url(self):
        return self.url
    
class Rakuten:
    def __init__(self,api,params):
        self.api = api
        self.params = params

    
    def get_requests(self,url,parameters):
        self.req = requests.get(url,params=parameters)
        return self.req

    def check_header(self,req):
        if "json" in req.headers.get("Content-type"):
            result = req.json()

            return result
        else:
            result = req.text
            return True
        
class RakutenAnalysis:
    def sampling(self,q:dict,keyword:str):
        """
        取得したAPI情報から
        インスタンス化する商品情報を選定

        Args:
            q (str): キーワード
        return:
            items(dict): 出力する商品情報
        """
        
        # 取得する商品情報を抽出
        items = q["Items"]
        items = [item["Item"] for item in items]
            # データを逐次追加する
        df = pd.DataFrame(items,columns=[
            'itemCode','itemName','itemPrice',
            'itemUrl','mediumImageUrls',
            'genreId','asurakuArea','availability',
            'catchcopy','reviewAverage', 'reviewCount',
            'shopCode', 'shopName','shopUrl'
        ])
        new_columns = ['商品番号','商品名','商品金額','商品URL','商品画像URL','ジャンルID','あす楽エリア','販売可能','キャッチコピー','レビュー平均点','レビュー数','ショップ番号','ショップ名','ショップURL']
        df.columns = new_columns
        df_s = df.sort_values('商品金額',ascending=False)#昇順ではない設定

        # pandasでexcelファイルを作成
        now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        df_s.to_excel(f"export_{keyword}_{now}.xlsx",sheet_name="商品情報")
        # CSVファイルを作成
        df_s.to_csv(f"export_{keyword}_{now}.csv",mode='w',encoding='utf-8 sig')
        wb = px.load_workbook(f"export_{keyword}_{now}.xlsx")
        sheet = wb["商品情報"]
        max = sheet.max_row
        min = sheet.min_row
        max = sheet.cell(row=max,column=4).value# 最終行
        min = sheet.cell(row=min+1,column=4).value#先頭行+1
        print('最終行の商品金額',max)
        print('先頭行の商品金額',min)

def main(keyword:str):
    # パラメータセット
    parameters = {
                'applicationId': '1090101428128448381', #アプリID
                'affiliateId' : '24e554d6.nc5b4d4a8.24e554d7.493927e4', #アフィリエイトID
                'keyword' : 'keyword', # キーワードの入力
                'format' : 'json'
    }
    url = "https://app.rakuten.co.jp/services/api/IchibaItem/Search/20170706"
    # 検索するキーワード取得

    if 'keyword' in parameters:
        parameters['keyword'] = keyword

    api = API(url)
    rakuten = Rakuten(api,parameters)
    # APIのリクエストを取得
    req = rakuten.get_requests(api.get_url(),parameters)

    # ヘッダー情報を確認する
    result = rakuten.check_header(req)
    # 結果からデータを抽出
    rakuten_analysis = RakutenAnalysis()
    
    rakuten_analysis.sampling(result,keyword)


if __name__ == "__main__":
    fire.Fire(main)